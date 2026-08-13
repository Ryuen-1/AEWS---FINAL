import openpyxl
from pathlib import Path

file_path = Path(r"C:\Users\Ian\Desktop\SYSTEM FOR 2ND SEM\stone\backend\uploads\Gradesheet Sample.xlsx")

print("=" * 80)
print("ANALYZING GRADESHEET FILE")
print("=" * 80)

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"\nWorkbook loaded successfully")
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        print('='*80)
        
        # Print first 50 rows
        for row_idx in range(1, min(51, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            # Print row if it has content
            if any(cell.strip() for cell in row_data):
                print(f"Row {row_idx}: {' | '.join(row_data[:10])}")  # Show first 10 columns
        
        print(f"\n... (total rows: {ws.max_row})")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
